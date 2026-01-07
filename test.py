"""
GymChat Pro - Premium Multi-Unit Gym Management System
Modern UI with glassmorphism, animations, and premium design

To create EXE:
pip install pyinstaller tkinter pillow python-dateutil openpyxl
pyinstaller --onefile --windowed --name GymChatPro gym_chat_pro.py
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import sqlite3
import pathlib
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import hashlib
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ModernStyle:
    """Modern UI styling with premium colors and effects"""
    
    # Premium color palette - Dark mode with vibrant accents
    BG_DARK = '#0A0E27'
    BG_CARD = '#1A1F3A'
    BG_CARD_HOVER = '#252B4A'
    
    ACCENT_PRIMARY = '#6C5CE7'  # Purple
    ACCENT_SECONDARY = '#A29BFE'  # Light purple
    ACCENT_SUCCESS = '#00D9A3'  # Mint green
    ACCENT_WARNING = '#FDCB6E'  # Gold
    ACCENT_DANGER = '#FF6B6B'  # Coral red
    
    TEXT_PRIMARY = '#FFFFFF'
    TEXT_SECONDARY = '#A0AEC0'
    TEXT_MUTED = '#718096'
    
    GRADIENT_START = '#667EEA'
    GRADIENT_END = '#764BA2'
    
    # Chat colors
    CHAT_BG = '#0F1419'
    MY_MESSAGE = '#6C5CE7'
    OTHER_MESSAGE = '#2D3748'
    
    # Fonts
    FONT_HEADING = ('Segoe UI', 24, 'bold')
    FONT_SUBHEADING = ('Segoe UI', 16, 'bold')
    FONT_BODY = ('Segoe UI', 11)
    FONT_BODY_BOLD = ('Segoe UI', 11, 'bold')
    FONT_SMALL = ('Segoe UI', 9)
    FONT_BUTTON = ('Segoe UI', 10, 'bold')

class DatabaseManager:
    """Manages multiple databases for different gym units"""
    
    def __init__(self):
        self.base_path = pathlib.Path('gym_units')
        self.base_path.mkdir(exist_ok=True)
        self.units = ['unit1', 'unit2', 'unit3', 'unit4', 'unit5', 'unit6', 'unit7']
        
    def get_db_path(self, unit):
        return self.base_path / f'{unit}.db'
    
    def get_db(self, unit):
        con = sqlite3.connect(self.get_db_path(unit))
        con.row_factory = sqlite3.Row
        return con
    
    def init_all_databases(self):
        for unit in self.units:
            self.init_unit_db(unit)
        self.init_auth_db()
    
    def init_unit_db(self, unit):
        con = self.get_db(unit)
        con.executescript("""
            PRAGMA foreign_keys = ON;
            
            CREATE TABLE IF NOT EXISTS members (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                phone TEXT NOT NULL UNIQUE,
                email TEXT,
                gender TEXT,
                dob TEXT,
                created_at TEXT DEFAULT (datetime('now')),
                avatar TEXT DEFAULT '👤'
            );
            
            CREATE TABLE IF NOT EXISTS plans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                duration_months INTEGER NOT NULL CHECK(duration_months > 0),
                price REAL NOT NULL CHECK(price >= 0)
            );
            
            CREATE TABLE IF NOT EXISTS subscriptions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                member_id INTEGER NOT NULL REFERENCES members(id) ON DELETE CASCADE,
                plan_id INTEGER NOT NULL REFERENCES plans(id),
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'active',
                notes TEXT
            );
            
            CREATE TABLE IF NOT EXISTS chats (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                chat_id TEXT UNIQUE NOT NULL,
                chat_name TEXT NOT NULL,
                chat_type TEXT NOT NULL,
                member_id INTEGER REFERENCES members(id),
                created_at TEXT DEFAULT (datetime('now'))
            );
            
            CREATE TABLE IF NOT EXISTS messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                chat_id TEXT NOT NULL,
                sender_name TEXT NOT NULL,
                sender_role TEXT NOT NULL,
                message TEXT NOT NULL,
                timestamp TEXT DEFAULT (datetime('now')),
                read_status INTEGER DEFAULT 0
            );
            
            CREATE TABLE IF NOT EXISTS notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                subscription_id INTEGER REFERENCES subscriptions(id),
                member_id INTEGER REFERENCES members(id),
                channel TEXT NOT NULL,
                message TEXT NOT NULL,
                sent_at TEXT,
                status TEXT DEFAULT 'pending'
            );
            
            INSERT OR IGNORE INTO plans(name, duration_months, price) VALUES 
                ('Monthly', 1, 999.0),
                ('Quarterly', 3, 2699.0),
                ('Half-Yearly', 6, 4999.0),
                ('Yearly', 12, 8999.0);
                
            INSERT OR IGNORE INTO chats(chat_id, chat_name, chat_type) 
            VALUES ('group_general', 'General Announcements', 'group');
        """)
        con.commit()
        
        # Add sample data if no members exist
        member_count = con.execute("SELECT COUNT(*) FROM members").fetchone()[0]
        if member_count == 0:
            self.add_sample_data(con)
        
        con.close()
    
    def add_sample_data(self, con):
        """Add sample members and subscriptions"""
        import random
        
        # Sample data
        avatars = ['👨', '👩', '🧑', '👨‍💼', '👩‍💼', '🧔', '👨‍🦱', '👩‍🦰', '👨‍🦳', '👩‍🦳']
        first_names = ['Raj', 'Priya', 'Amit', 'Sneha', 'Vikram', 'Anjali', 'Rahul', 'Divya', 
                      'Arjun', 'Kavya', 'Karthik', 'Meera', 'Suresh', 'Lakshmi', 'Arun']
        last_names = ['Kumar', 'Sharma', 'Patel', 'Singh', 'Reddy', 'Nair', 'Iyer', 'Gupta',
                     'Rao', 'Joshi', 'Mehta', 'Shah', 'Pillai', 'Menon', 'Agarwal']
        
        today = date.today()
        
        # Reduced to 10 sample members for faster initial load
        members = []
        for i in range(10):
            name = f"{random.choice(first_names)} {random.choice(last_names)}"
            phone = f"+91{7000000000 + i * 111111111}"  # Sequential to avoid collisions
            email = f"member{i}@email.com"
            gender = random.choice(['Male', 'Female'])
            avatar = random.choice(avatars)
            dob = date(random.randint(1985, 2000), random.randint(1, 12), random.randint(1, 28)).isoformat()
            
            cur = con.execute("""
                INSERT INTO members(name, phone, email, gender, dob, avatar)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (name, phone, email, gender, dob, avatar))
            members.append(cur.lastrowid)
        
        # Create subscriptions with various expiry dates
        plan_ids = [1, 2, 3, 4]  # Monthly, Quarterly, Half-Yearly, Yearly
        
        # Batch insert for better performance
        subscription_data = []
        for member_id in members:
            plan_id = random.choice(plan_ids)
            
            # Get plan duration
            plan = con.execute("SELECT duration_months FROM plans WHERE id = ?", (plan_id,)).fetchone()
            duration = plan[0]
            
            # Random start date in past (0-6 months ago)
            months_ago = random.randint(0, 6)
            start_date = today - relativedelta(months=months_ago)
            end_date = start_date + relativedelta(months=duration)
            
            # Create different scenarios
            scenario = random.random()
            if scenario < 0.2:  # 20% expired
                start_date = today - relativedelta(months=duration+2)
                end_date = today - relativedelta(months=2)
            elif scenario < 0.4:  # 20% expiring within 30 days
                start_date = today - relativedelta(months=duration, days=-15)
                end_date = today + relativedelta(days=15)
            
            notes = random.choice(['Regular member', 'Morning batch', 'Evening batch', ''])
            
            subscription_data.append((member_id, plan_id, start_date.isoformat(), 
                                    end_date.isoformat(), notes))
        
        # Batch insert all subscriptions
        con.executemany("""
            INSERT INTO subscriptions(member_id, plan_id, start_date, end_date, status, notes)
            VALUES (?, ?, ?, ?, 'active', ?)
        """, subscription_data)
        
        con.commit()
    
    def init_auth_db(self):
        auth_path = self.base_path / 'auth.db'
        con = sqlite3.connect(auth_path)
        con.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL,
                unit TEXT,
                full_name TEXT NOT NULL,
                created_at TEXT DEFAULT (datetime('now'))
            );
            
            CREATE TABLE IF NOT EXISTS login_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER REFERENCES users(id),
                login_time TEXT DEFAULT (datetime('now')),
                unit_accessed TEXT
            );
        """)
        
        owner_hash = hashlib.sha256('owner123'.encode()).hexdigest()
        con.execute("""
            INSERT OR IGNORE INTO users(username, password_hash, role, full_name)
            VALUES ('owner', ?, 'owner', 'Gym Owner')
        """, (owner_hash,))
        
        for unit in self.units:
            unit_hash = hashlib.sha256(unit.encode()).hexdigest()
            con.execute("""
                INSERT OR IGNORE INTO users(username, password_hash, role, unit, full_name)
                VALUES (?, ?, 'unit_admin', ?, ?)
            """, (unit, unit_hash, unit, f'{unit.title()} Admin'))
        
        con.commit()
        con.close()

class ModernButton(tk.Canvas):
    """Custom modern button with hover effects"""
    
    def __init__(self, parent, text, command, bg_color, fg_color='white', 
                 width=150, height=45, icon='', **kwargs):
        super().__init__(parent, width=width, height=height, 
                        bg=parent['bg'], highlightthickness=0, **kwargs)
        
        self.bg_color = bg_color
        self.fg_color = fg_color
        self.text = text
        self.icon = icon
        self.command = command
        self.hover = False
        
        self.draw_button()
        self.bind('<Button-1>', lambda e: self.on_click())
        self.bind('<Enter>', lambda e: self.on_hover(True))
        self.bind('<Leave>', lambda e: self.on_hover(False))
        
    def draw_button(self):
        self.delete('all')
        
        # Button background with rounded corners
        radius = 23
        color = self.adjust_color(self.bg_color, 1.2 if self.hover else 1.0)
        
        self.create_arc(0, 0, radius*2, radius*2, 
                       start=90, extent=90, fill=color, outline='')
        self.create_arc(self.winfo_reqwidth()-radius*2, 0, 
                       self.winfo_reqwidth(), radius*2,
                       start=0, extent=90, fill=color, outline='')
        self.create_arc(0, self.winfo_reqheight()-radius*2, 
                       radius*2, self.winfo_reqheight(),
                       start=180, extent=90, fill=color, outline='')
        self.create_arc(self.winfo_reqwidth()-radius*2, 
                       self.winfo_reqheight()-radius*2,
                       self.winfo_reqwidth(), self.winfo_reqheight(),
                       start=270, extent=90, fill=color, outline='')
        
        self.create_rectangle(radius, 0, self.winfo_reqwidth()-radius, 
                            self.winfo_reqheight(), fill=color, outline='')
        self.create_rectangle(0, radius, self.winfo_reqwidth(), 
                            self.winfo_reqheight()-radius, fill=color, outline='')
        
        # Text
        full_text = f"{self.icon} {self.text}" if self.icon else self.text
        self.create_text(self.winfo_reqwidth()//2, self.winfo_reqheight()//2,
                        text=full_text, fill=self.fg_color, 
                        font=ModernStyle.FONT_BUTTON)
    
    def adjust_color(self, color, factor):
        """Lighten or darken a color"""
        color = color.lstrip('#')
        r, g, b = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
        r = min(255, int(r * factor))
        g = min(255, int(g * factor))
        b = min(255, int(b * factor))
        return f'#{r:02x}{g:02x}{b:02x}'
    
    def on_hover(self, entering):
        self.hover = entering
        self.draw_button()
        self.config(cursor='hand2' if entering else '')
    
    def on_click(self):
        if self.command:
            self.command()

class ModernEntry(tk.Frame):
    """Modern entry field with floating label effect"""
    
    def __init__(self, parent, label, show=None):
        super().__init__(parent, bg=ModernStyle.BG_CARD)
        
        self.label_text = label
        
        tk.Label(self, text=label, font=ModernStyle.FONT_SMALL,
                bg=ModernStyle.BG_CARD, fg=ModernStyle.TEXT_SECONDARY).pack(anchor='w', pady=(0, 5))
        
        entry_frame = tk.Frame(self, bg=ModernStyle.BG_DARK, bd=0)
        entry_frame.pack(fill=tk.X)
        
        self.entry = tk.Entry(entry_frame, font=ModernStyle.FONT_BODY,
                             bg=ModernStyle.BG_DARK, fg=ModernStyle.TEXT_PRIMARY,
                             insertbackground=ModernStyle.ACCENT_PRIMARY,
                             relief=tk.FLAT, bd=0, show=show)
        self.entry.pack(fill=tk.X, padx=15, pady=12)
        
    def get(self):
        return self.entry.get()
    
    def insert(self, index, text):
        self.entry.insert(index, text)

class LoginWindow:
    """Premium login window with gradient background"""
    
    def __init__(self, parent, on_success):
        self.parent = parent
        self.on_success = on_success
        self.window = tk.Toplevel(parent)
        self.window.title("GymChat Pro - Login")
        self.window.geometry("500x700")
        self.window.configure(bg=ModernStyle.BG_DARK)
        self.window.transient(parent)
        self.window.grab_set()
        
        # Center window
        self.window.update_idletasks()
        x = (self.window.winfo_screenwidth() // 2) - 250
        y = (self.window.winfo_screenheight() // 2) - 350
        self.window.geometry(f'500x700+{x}+{y}')
        
        self.setup_ui()
        
    def setup_ui(self):
        # Header with gradient effect
        header = tk.Frame(self.window, bg=ModernStyle.BG_CARD, height=200)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        # Logo and title
        logo_frame = tk.Frame(header, bg=ModernStyle.BG_CARD)
        logo_frame.pack(expand=True)
        
        tk.Label(logo_frame, text="💪", font=('Arial', 60),
                bg=ModernStyle.BG_CARD).pack()
        
        tk.Label(logo_frame, text="GymChat Pro", 
                font=('Segoe UI', 32, 'bold'),
                bg=ModernStyle.BG_CARD, 
                fg=ModernStyle.TEXT_PRIMARY).pack()
        
        tk.Label(logo_frame, text="Multi-Unit Management System", 
                font=ModernStyle.FONT_BODY,
                bg=ModernStyle.BG_CARD, 
                fg=ModernStyle.TEXT_SECONDARY).pack(pady=(5, 0))
        
        # Form container
        form_container = tk.Frame(self.window, bg=ModernStyle.BG_DARK)
        form_container.pack(fill=tk.BOTH, expand=True, padx=40, pady=40)
        
        # Login form card
        form_card = tk.Frame(form_container, bg=ModernStyle.BG_CARD, bd=0)
        form_card.pack(fill=tk.BOTH, expand=True)
        
        form_inner = tk.Frame(form_card, bg=ModernStyle.BG_CARD)
        form_inner.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)
        
        tk.Label(form_inner, text="Welcome Back!", 
                font=ModernStyle.FONT_SUBHEADING,
                bg=ModernStyle.BG_CARD, 
                fg=ModernStyle.TEXT_PRIMARY).pack(pady=(0, 30))
        
        # Username field
        self.username_field = ModernEntry(form_inner, "Username")
        self.username_field.pack(fill=tk.X, pady=(0, 20))
        
        # Password field
        self.password_field = ModernEntry(form_inner, "Password", show='●')
        self.password_field.pack(fill=tk.X, pady=(0, 30))
        self.password_field.entry.bind('<Return>', lambda e: self.login())
        
        # Login button
        btn_frame = tk.Frame(form_inner, bg=ModernStyle.BG_CARD)
        btn_frame.pack(fill=tk.X)
        
        login_btn = ModernButton(btn_frame, "Login", self.login,
                                ModernStyle.ACCENT_PRIMARY, 
                                width=400, height=50, icon='🔐')
        login_btn.pack(pady=(0, 20))
        
        # Info card
        info_card = tk.Frame(form_inner, bg=ModernStyle.BG_DARK, bd=0)
        info_card.pack(fill=tk.X, pady=(10, 0))
        
        info_inner = tk.Frame(info_card, bg=ModernStyle.BG_DARK)
        info_inner.pack(fill=tk.X, padx=20, pady=20)
        
        tk.Label(info_inner, text="ℹ  Default Credentials", 
                font=ModernStyle.FONT_BODY_BOLD,
                bg=ModernStyle.BG_DARK, 
                fg=ModernStyle.ACCENT_WARNING).pack(pady=(0, 10))
        
        tk.Label(info_inner, text="Owner: owner / owner123",
                font=ModernStyle.FONT_SMALL, bg=ModernStyle.BG_DARK,
                fg=ModernStyle.TEXT_SECONDARY).pack(anchor='w', padx=20)
        
        tk.Label(info_inner, text="Unit Admin: unit1 / unit1",
                font=ModernStyle.FONT_SMALL, bg=ModernStyle.BG_DARK,
                fg=ModernStyle.TEXT_SECONDARY).pack(anchor='w', padx=20)
        
        tk.Label(info_inner, text="(unit2-unit7 follow same pattern)",
                font=ModernStyle.FONT_SMALL, bg=ModernStyle.BG_DARK,
                fg=ModernStyle.TEXT_MUTED).pack(anchor='w', padx=20, pady=(0, 5))
        
    def login(self):
        username = self.username_field.get().strip()
        password = self.password_field.get().strip()
        
        if not username or not password:
            messagebox.showerror("Error", "Please enter username and password")
            return
        
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        auth_path = pathlib.Path('gym_units') / 'auth.db'
        con = sqlite3.connect(auth_path)
        con.row_factory = sqlite3.Row
        
        user = con.execute("""
            SELECT * FROM users 
            WHERE username = ? AND password_hash = ?
        """, (username, password_hash)).fetchone()
        
        if user:
            con.execute("""
                INSERT INTO login_logs(user_id, unit_accessed)
                VALUES (?, ?)
            """, (user['id'], user['unit']))
            con.commit()
            con.close()
            
            user_data = {
                'id': user['id'],
                'username': user['username'],
                'role': user['role'],
                'unit': user['unit'],
                'full_name': user['full_name']
            }
            
            self.window.destroy()
            self.on_success(user_data)
        else:
            con.close()
            messagebox.showerror("Login Failed", "Invalid username or password")

class GymChatApp:
    """Main application with premium UI"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("GymChat Pro - Multi-Unit Management")
        self.root.geometry("1400x850")
        self.root.minsize(1200, 700)
        self.root.configure(bg=ModernStyle.BG_DARK)
        
        # Show loading splash
        self.show_loading_splash()
        
        # Initialize database in background
        self.root.after(100, self.init_database)
        
    def show_loading_splash(self):
        """Show loading splash screen"""
        self.splash = tk.Toplevel(self.root)
        self.splash.title("Loading...")
        self.splash.geometry("400x300")
        self.splash.configure(bg=ModernStyle.BG_CARD)
        self.splash.overrideredirect(True)
        
        # Center splash
        self.splash.update_idletasks()
        x = (self.splash.winfo_screenwidth() // 2) - 200
        y = (self.splash.winfo_screenheight() // 2) - 150
        self.splash.geometry(f'400x300+{x}+{y}')
        
        tk.Label(self.splash, text="💪", font=('Arial', 60),
                bg=ModernStyle.BG_CARD).pack(pady=(40, 20))
        
        tk.Label(self.splash, text="GymChat Pro", 
                font=('Segoe UI', 28, 'bold'),
                bg=ModernStyle.BG_CARD, 
                fg=ModernStyle.TEXT_PRIMARY).pack()
        
        tk.Label(self.splash, text="Loading databases...", 
                font=ModernStyle.FONT_BODY,
                bg=ModernStyle.BG_CARD, 
                fg=ModernStyle.TEXT_SECONDARY).pack(pady=(20, 0))
        
    def init_database(self):
        """Initialize database"""
        self.db_manager = DatabaseManager()
        self.db_manager.init_all_databases()
        
        self.current_user = None
        self.current_unit = None
        self.active_chat = None
        
        # Close splash and show login
        self.splash.destroy()
        self.root.withdraw()
        LoginWindow(self.root, self.on_login_success)
        
    def on_login_success(self, user_data):
        self.current_user = user_data
        
        if user_data['role'] == 'owner':
            self.accessible_units = self.db_manager.units
            self.current_unit = 'unit1'
        else:
            self.accessible_units = [user_data['unit']]
            self.current_unit = user_data['unit']
        
        self.root.deiconify()
        self.setup_ui()
        
    def create_gradient_header(self, parent):
        """Create gradient header"""
        canvas = tk.Canvas(parent, height=80, bg=ModernStyle.BG_CARD, 
                          highlightthickness=0)
        canvas.pack(fill=tk.X)
        
        # Create gradient effect using overlapping rectangles
        for i in range(80):
            ratio = i / 80
            r1, g1, b1 = int('66', 16), int('7E', 16), int('EA', 16)
            r2, g2, b2 = int('76', 16), int('4B', 16), int('A2', 16)
            
            r = int(r1 + (r2 - r1) * ratio)
            g = int(g1 + (g2 - g1) * ratio)
            b = int(b1 + (b2 - b1) * ratio)
            
            color = f'#{r:02x}{g:02x}{b:02x}'
            canvas.create_line(0, i, 1400, i, fill=color)
        
        return canvas
        
    def setup_ui(self):
        # Premium top bar with gradient
        top_bar = tk.Frame(self.root, height=80)
        top_bar.pack(fill=tk.X, side=tk.TOP)
        top_bar.pack_propagate(False)
        
        gradient_canvas = self.create_gradient_header(top_bar)
        
        # Title and logo
        title_frame = tk.Frame(gradient_canvas, bg='')
        gradient_canvas.create_window(20, 40, window=title_frame, anchor='w')
        
        tk.Label(title_frame, text="💪", font=('Arial', 28),
                bg='').pack(side=tk.LEFT, padx=(0, 10))
        
        title_text = tk.Frame(title_frame, bg='')
        title_text.pack(side=tk.LEFT)
        
        tk.Label(title_text, text="GymChat Pro", 
                font=('Segoe UI', 20, 'bold'),
                fg=ModernStyle.TEXT_PRIMARY, bg='').pack(anchor='w')
        
        tk.Label(title_text, text="Multi-Unit Management System", 
                font=ModernStyle.FONT_SMALL,
                fg=ModernStyle.TEXT_SECONDARY, bg='').pack(anchor='w')
        
        # Unit selector for owner
        if self.current_user['role'] == 'owner':
            unit_frame = tk.Frame(gradient_canvas, bg='')
            gradient_canvas.create_window(500, 40, window=unit_frame, anchor='w')
            
            tk.Label(unit_frame, text="Active Unit:", 
                    font=ModernStyle.FONT_BODY_BOLD,
                    fg=ModernStyle.TEXT_PRIMARY, bg='').pack(side=tk.LEFT, padx=(0, 10))
            
            style = ttk.Style()
            style.configure('Unit.TCombobox', 
                          fieldbackground=ModernStyle.BG_DARK,
                          background=ModernStyle.ACCENT_PRIMARY,
                          foreground=ModernStyle.TEXT_PRIMARY)
            
            self.unit_var = tk.StringVar(value=self.current_unit)
            unit_menu = ttk.Combobox(unit_frame, textvariable=self.unit_var,
                                    values=self.accessible_units, state='readonly',
                                    width=12, font=ModernStyle.FONT_BODY,
                                    style='Unit.TCombobox')
            unit_menu.pack(side=tk.LEFT)
            unit_menu.bind('<<ComboboxSelected>>', self.on_unit_change)
        
        # User profile section
        user_frame = tk.Frame(gradient_canvas, bg='')
        gradient_canvas.create_window(1350, 40, window=user_frame, anchor='e')
        
        profile_card = tk.Frame(user_frame, bg=ModernStyle.BG_CARD_HOVER, bd=0)
        profile_card.pack(side=tk.RIGHT)
        
        profile_inner = tk.Frame(profile_card, bg=ModernStyle.BG_CARD_HOVER)
        profile_inner.pack(padx=15, pady=8)
        
        role_emoji = '👑' if self.current_user['role'] == 'owner' else '⭐'
        
        tk.Label(profile_inner, text=f"{role_emoji} {self.current_user['full_name']}", 
                font=ModernStyle.FONT_BODY_BOLD,
                bg=ModernStyle.BG_CARD_HOVER, 
                fg=ModernStyle.TEXT_PRIMARY).pack(side=tk.LEFT, padx=(0, 15))
        
        logout_btn = tk.Label(profile_inner, text="🚪 Logout",
                            font=ModernStyle.FONT_SMALL,
                            bg=ModernStyle.ACCENT_DANGER, 
                            fg='white', cursor='hand2',
                            padx=12, pady=4)
        logout_btn.pack(side=tk.LEFT)
        logout_btn.bind('<Button-1>', lambda e: self.logout())
        
        # Main container
        main_container = tk.Frame(self.root, bg=ModernStyle.BG_DARK)
        main_container.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        
        self.setup_content_area(main_container)
        
    def setup_content_area(self, parent):
        content = tk.Frame(parent, bg=ModernStyle.BG_DARK)
        content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title
        title_frame = tk.Frame(content, bg=ModernStyle.BG_DARK)
        title_frame.pack(fill=tk.X, pady=(0, 20))
        
        tk.Label(title_frame, text=f"📊 Membership Manager - {self.current_unit.upper()}", 
                font=ModernStyle.FONT_HEADING,
                bg=ModernStyle.BG_DARK, 
                fg=ModernStyle.TEXT_PRIMARY).pack(side=tk.LEFT)
        
        # Month filter
        filter_frame = tk.Frame(content, bg=ModernStyle.BG_CARD)
        filter_frame.pack(fill=tk.X, pady=(0, 20))
        
        filter_inner = tk.Frame(filter_frame, bg=ModernStyle.BG_CARD)
        filter_inner.pack(padx=20, pady=15)
        
        tk.Label(filter_inner, text="Filter by Expiry Month:", 
                font=ModernStyle.FONT_BODY_BOLD,
                bg=ModernStyle.BG_CARD, 
                fg=ModernStyle.TEXT_PRIMARY).pack(side=tk.LEFT, padx=(0, 15))
        
        # Month selector
        months = ['All', 'January', 'February', 'March', 'April', 'May', 'June',
                 'July', 'August', 'September', 'October', 'November', 'December']
        
        self.month_var = tk.StringVar(value='All')
        month_menu = ttk.Combobox(filter_inner, textvariable=self.month_var,
                                 values=months, state='readonly',
                                 width=15, font=ModernStyle.FONT_BODY)
        month_menu.pack(side=tk.LEFT, padx=(0, 15))
        month_menu.bind('<<ComboboxSelected>>', lambda e: self.load_memberships())
        
        # Year selector
        current_year = datetime.now().year
        years = ['All'] + [str(y) for y in range(current_year, current_year + 3)]
        
        self.year_var = tk.StringVar(value=str(current_year))
        year_menu = ttk.Combobox(filter_inner, textvariable=self.year_var,
                                values=years, state='readonly',
                                width=10, font=ModernStyle.FONT_BODY)
        year_menu.pack(side=tk.LEFT, padx=(0, 15))
        year_menu.bind('<<ComboboxSelected>>', lambda e: self.load_memberships())
        
        # Export button
        export_btn = ModernButton(filter_inner, "Export to Excel", self.export_to_excel,
                                 ModernStyle.ACCENT_SUCCESS, width=180, height=40, icon='📥')
        export_btn.pack(side=tk.LEFT, padx=(15, 0))
        
        # Membership table
        table_frame = tk.Frame(content, bg=ModernStyle.BG_CARD)
        table_frame.pack(fill=tk.BOTH, expand=True)
        
        # Table header
        header_frame = tk.Frame(table_frame, bg=ModernStyle.ACCENT_PRIMARY, height=45)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        columns = [
            ("Member Name", 200),
            ("Phone", 150),
            ("Plan", 120),
            ("Start Date", 120),
            ("End Date", 120),
            ("Days Left", 100),
            ("Status", 120)
        ]
        
        for col_name, width in columns:
            tk.Label(header_frame, text=col_name, 
                    font=ModernStyle.FONT_BODY_BOLD,
                    bg=ModernStyle.ACCENT_PRIMARY, 
                    fg='white', width=width//8).pack(side=tk.LEFT, padx=5, pady=10)
        
        # Scrollable content
        canvas = tk.Canvas(table_frame, bg=ModernStyle.BG_CARD, highlightthickness=0)
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=canvas.yview)
        self.membership_frame = tk.Frame(canvas, bg=ModernStyle.BG_CARD)
        
        self.membership_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.membership_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Load initial data
        self.load_memberships()
        
    def load_memberships(self):
        """Load membership data based on filters"""
        # Clear existing rows
        for widget in self.membership_frame.winfo_children():
            widget.destroy()
        
        con = self.db_manager.get_db(self.current_unit)
        
        # Build query based on filters
        query = """
            SELECT 
                m.id, m.name, m.phone, m.avatar,
                p.name as plan_name,
                s.start_date, s.end_date, s.status
            FROM subscriptions s
            JOIN members m ON s.member_id = m.id
            JOIN plans p ON s.plan_id = p.id
            WHERE s.status = 'active'
        """
        
        params = []
        
        # Apply month filter
        if self.month_var.get() != 'All':
            month_num = ['January', 'February', 'March', 'April', 'May', 'June',
                        'July', 'August', 'September', 'October', 'November', 'December'].index(self.month_var.get()) + 1
            query += " AND CAST(strftime('%m', s.end_date) AS INTEGER) = ?"
            params.append(month_num)
        
        # Apply year filter
        if self.year_var.get() != 'All':
            query += " AND strftime('%Y', s.end_date) = ?"
            params.append(self.year_var.get())
        
        query += " ORDER BY s.end_date ASC"
        
        memberships = con.execute(query, params).fetchall()
        con.close()
        
        today = date.today()
        
        for idx, membership in enumerate(memberships):
            end_date = datetime.strptime(membership['end_date'], '%Y-%m-%d').date()
            days_left = (end_date - today).days
            
            # Determine status color
            if days_left < 0:
                status_text = "Expired"
                status_color = ModernStyle.ACCENT_DANGER
            elif days_left <= 7:
                status_text = "Expiring Soon"
                status_color = ModernStyle.ACCENT_DANGER
            elif days_left <= 30:
                status_text = "Expiring"
                status_color = ModernStyle.ACCENT_WARNING
            else:
                status_text = "Active"
                status_color = ModernStyle.ACCENT_SUCCESS
            
            # Row frame
            row_bg = ModernStyle.BG_DARK if idx % 2 == 0 else ModernStyle.BG_CARD_HOVER
            row_frame = tk.Frame(self.membership_frame, bg=row_bg, height=50)
            row_frame.pack(fill=tk.X)
            row_frame.pack_propagate(False)
            
            # Member name with avatar
            name_frame = tk.Frame(row_frame, bg=row_bg, width=200)
            name_frame.pack(side=tk.LEFT, padx=5, pady=5)
            name_frame.pack_propagate(False)
            
            tk.Label(name_frame, text=f"{membership['avatar']} {membership['name']}", 
                    font=ModernStyle.FONT_BODY,
                    bg=row_bg, fg=ModernStyle.TEXT_PRIMARY,
                    anchor='w').pack(fill=tk.X, padx=5)
            
            # Phone
            phone_frame = tk.Frame(row_frame, bg=row_bg, width=150)
            phone_frame.pack(side=tk.LEFT, padx=5)
            phone_frame.pack_propagate(False)
            
            tk.Label(phone_frame, text=membership['phone'], 
                    font=ModernStyle.FONT_BODY,
                    bg=row_bg, fg=ModernStyle.TEXT_SECONDARY,
                    anchor='w').pack(fill=tk.X, padx=5)
            
            # Plan
            plan_frame = tk.Frame(row_frame, bg=row_bg, width=120)
            plan_frame.pack(side=tk.LEFT, padx=5)
            plan_frame.pack_propagate(False)
            
            tk.Label(plan_frame, text=membership['plan_name'], 
                    font=ModernStyle.FONT_BODY_BOLD,
                    bg=row_bg, fg=ModernStyle.ACCENT_SECONDARY,
                    anchor='w').pack(fill=tk.X, padx=5)
            
            # Start date
            start_frame = tk.Frame(row_frame, bg=row_bg, width=120)
            start_frame.pack(side=tk.LEFT, padx=5)
            start_frame.pack_propagate(False)
            
            tk.Label(start_frame, text=membership['start_date'], 
                    font=ModernStyle.FONT_BODY,
                    bg=row_bg, fg=ModernStyle.TEXT_SECONDARY,
                    anchor='w').pack(fill=tk.X, padx=5)
            
            # End date
            end_frame = tk.Frame(row_frame, bg=row_bg, width=120)
            end_frame.pack(side=tk.LEFT, padx=5)
            end_frame.pack_propagate(False)
            
            tk.Label(end_frame, text=membership['end_date'], 
                    font=ModernStyle.FONT_BODY,
                    bg=row_bg, fg=ModernStyle.TEXT_SECONDARY,
                    anchor='w').pack(fill=tk.X, padx=5)
            
            # Days left
            days_frame = tk.Frame(row_frame, bg=row_bg, width=100)
            days_frame.pack(side=tk.LEFT, padx=5)
            days_frame.pack_propagate(False)
            
            days_text = str(days_left) if days_left >= 0 else "Expired"
            tk.Label(days_frame, text=days_text, 
                    font=ModernStyle.FONT_BODY_BOLD,
                    bg=row_bg, fg=status_color,
                    anchor='center').pack(fill=tk.X, padx=5)
            
            # Status badge
            status_frame = tk.Frame(row_frame, bg=row_bg, width=120)
            status_frame.pack(side=tk.LEFT, padx=5)
            status_frame.pack_propagate(False)
            
            status_badge = tk.Label(status_frame, text=status_text, 
                                   font=ModernStyle.FONT_SMALL,
                                   bg=status_color, fg='white',
                                   padx=10, pady=4)
            status_badge.pack(pady=8)
        
        # Show count
        count_label = tk.Label(self.membership_frame, 
                              text=f"\nTotal Members: {len(memberships)}", 
                              font=ModernStyle.FONT_BODY_BOLD,
                              bg=ModernStyle.BG_CARD, 
                              fg=ModernStyle.TEXT_PRIMARY)
        count_label.pack(pady=20)
    
    def export_to_excel(self):
        """Export membership data to Excel"""
        try:
            # Get filtered data
            con = self.db_manager.get_db(self.current_unit)
            
            query = """
                SELECT 
                    m.name, m.phone, m.email, m.gender,
                    p.name as plan_name, p.price,
                    s.start_date, s.end_date, s.status, s.notes
                FROM subscriptions s
                JOIN members m ON s.member_id = m.id
                JOIN plans p ON s.plan_id = p.id
                WHERE s.status = 'active'
            """
            
            params = []
            
            if self.month_var.get() != 'All':
                month_num = ['January', 'February', 'March', 'April', 'May', 'June',
                            'July', 'August', 'September', 'October', 'November', 'December'].index(self.month_var.get()) + 1
                query += " AND CAST(strftime('%m', s.end_date) AS INTEGER) = ?"
                params.append(month_num)
            
            if self.year_var.get() != 'All':
                query += " AND strftime('%Y', s.end_date) = ?"
                params.append(self.year_var.get())
            
            query += " ORDER BY s.end_date ASC"
            
            memberships = con.execute(query, params).fetchall()
            con.close()
            
            if not memberships:
                messagebox.showinfo("No Data", "No memberships found with current filters.")
                return
            
            # Ask user where to save
            filename = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel files', '*.xlsx')],
                initialfile=f'memberships_{self.current_unit}_{datetime.now().strftime("%Y%m%d")}.xlsx'
            )
            
            if not filename:
                return
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"{self.current_unit.upper()} Memberships"
            
            # Styling
            header_fill = PatternFill(start_color='6C5CE7', end_color='6C5CE7', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = ['Member Name', 'Phone', 'Email', 'Gender', 'Plan', 'Price', 
                      'Start Date', 'End Date', 'Days Left', 'Status', 'Notes']
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Data rows
            today = date.today()
            
            for row_num, membership in enumerate(memberships, 2):
                end_date = datetime.strptime(membership['end_date'], '%Y-%m-%d').date()
                days_left = (end_date - today).days
                
                if days_left < 0:
                    status_text = "Expired"
                    row_color = 'FF6B6B'
                elif days_left <= 7:
                    status_text = "Expiring Soon"
                    row_color = 'FF6B6B'
                elif days_left <= 30:
                    status_text = "Expiring"
                    row_color = 'FDCB6E'
                else:
                    status_text = "Active"
                    row_color = '00D9A3'
                
                row_data = [
                    membership['name'],
                    membership['phone'],
                    membership['email'] or 'N/A',
                    membership['gender'] or 'N/A',
                    membership['plan_name'],
                    f"₹{membership['price']:.2f}",
                    membership['start_date'],
                    membership['end_date'],
                    days_left if days_left >= 0 else 'Expired',
                    status_text,
                    membership['notes'] or ''
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Color status cell
                    if col_num == 10:  # Status column
                        cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adjust column widths
            column_widths = [20, 15, 25, 10, 15, 12, 12, 12, 12, 15, 30]
            for col_num, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col_num)].width = width
            
            # Add summary
            summary_row = len(memberships) + 3
            ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True, size=14)
            ws.cell(row=summary_row + 1, column=1, value=f"Total Members: {len(memberships)}")
            ws.cell(row=summary_row + 2, column=1, value=f"Unit: {self.current_unit.upper()}")
            ws.cell(row=summary_row + 3, column=1, value=f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            # Save workbook
            wb.save(filename)
            
            messagebox.showinfo("Success", f"Membership data exported successfully!\n\nFile saved: {filename}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export data:\n{str(e)}")
    
    def on_unit_change(self, event=None):
        """Handle unit change"""
        self.current_unit = self.unit_var.get()
        self.load_memberships()
    
    def logout(self):
        """Logout user"""
        if messagebox.askyesno("Logout", "Are you sure you want to logout?"):
            self.root.withdraw()
            for widget in self.root.winfo_children():
                widget.destroy()
            self.current_user = None
            self.current_unit = None
            LoginWindow(self.root, self.on_login_success)

if __name__ == '__main__':
    root = tk.Tk()
    app = GymChatApp(root)
    root.mainloop()